#!/usr/bin/env python3

import argparse
import datetime
from collections import defaultdict

import openpyxl


def read_excel_file(file_path: str):
    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Read the first row
    #first_row = []
    #for cell in sheet[1]:
    #    first_row.append(cell.value)
    #print("First Row:", first_row)

    # Read the first column
    #first_column = []
    #for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    #    for cell in row:
    #        first_column.append(cell.value)
    #print("First Column:", first_column)

    names = defaultdict(list)

    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=19)):
        if row_index <= 1: 
            continue
        for col_index, cell in enumerate(row):
            # ignore first two columns
            if col_index <= 1:
                continue
            when = sheet[1][col_index].value
            what = sheet[f'A{row_index+1}'].value
            what = sheet[f'A{row_index+1}'].value
            detail = sheet[f'B{row_index+1}'].value

            # print(f"{cell.value} = {when}: {what}")
            names[f"{cell.value}"].append((when, what, detail))

    print("<html>")
    print("""<head>
          <style>
          @media print }
            h1,h2 {
                page-break-after: avoid;
                page-break-before:auto;
            ul {
                break-inside: avoid-page;
                page-break-inside: avoid; 
                page-break-after:auto;
            }
          }
          </style>
    </head>""")
    print("<body>")

    print("<h1>Schichtplan VAWC, 2025</h1>")
    print()
    formatted_date_time = datetime.datetime.now().strftime('%d. %B %Y, %H:%M')
    print(f"<ul><li>Stand: {formatted_date_time} Uhr</ul></li>")
    print()
    sorted_names = sorted(names.keys())
    for name in sorted_names:
        if name is None or name == 'None' or name == "TBD" or name == "Aufbau" or name == 'Aufbau/Betrieb' or name == 'Betrieb':
            continue
        shifts = names[name]

        print(f'<h2>{name}</h2>')
        print()
        print("<ul>")
        for shift in shifts:
            if shift[2] is not None:
                print(f'<li>{shift[0]}: {shift[1]} - {shift[2]}</li>')
            else:
                print(f'<li>{shift[0]}: {shift[1]}</li>')
        print("</ul>")
    print("</body>")
    print("</html>")

def main(filename: str):
    read_excel_file(filename)


if __name__ == "__main__":
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Create a shift plan from an Excel Sheet for Quax e.V. events.')
    parser.add_argument('file_path', type=str, help='The path to the Excel file')

    # Parse the arguments
    args = parser.parse_args()

    # Call the function with the provided file path
    read_excel_file(args.file_path)
